# main.py
from  openpyxl import load_workbook
from xml.dom import minidom 
import lxml.etree as ET
import os  
import xlsxwriter

# from create_table import create_table
# from create_pdf import create_pdf

def main():
    wb = load_workbook("forras.xlsx", data_only=True)
    # table_data = create_table(excel_data)
    # create_pdf(table_data, "output.pdf")

    root = minidom.Document() 

    xml = root.createElement('iskola')  
    root.appendChild(xml) 

    productChild = root.createElement('product') 
    productChild.setAttribute('name', 'Geeks for Geeks') 
    xml.appendChild(productChild) 

    
    for sheetname in wb.sheetnames:
        if sheetname[0] in ['5','6','7','8']:

            sheet = wb[sheetname]
            lID = sheetname
            lName_orig = sheet.cell(row=1, column=1).value
            lName = lName_orig.split("2023/2024.")
        
            lName_Osztaly = lName[0].replace("-","").strip()
            lName_Lista = lName[1]
            lName_Lista = lName_Lista.split("-")
            if lName_Osztaly.strip() not in ['7.a','7.b']:
                lName_Tipus = lName_Lista[-2].strip()
                lName_Ofo = lName_Lista[-1].strip()
            else:
                lName_Tipus = lName_Lista[-3].strip()
                lName_Ofo = (lName_Lista[-1] + lName_Lista[-1]).strip()
            
            osztalyChild = root.createElement("osztaly") 
            osztalyChild.setAttribute('osztaly', lName_Osztaly) 
            osztalyChild.setAttribute('tipus', lName_Tipus) 
            osztalyChild.setAttribute('ofo', lName_Ofo) 
            xml.appendChild(osztalyChild) 

            row_num = 1
            while (sheet.cell(row = row_num, column= 1).value == None) or (sheet.cell(row = row_num, column= 1).value != 'Ssz.'):
                row_num += 1
            row_num += 1

            while row_num < 50 :
                lSorsz = ''
                lNev = ''
                lKoz = ''
                lOMAzon = ''
                lMegjegyzes = ''

                lSorsz = sheet.cell(row=row_num, column=1).value
                lNev = sheet.cell(row=row_num, column=2).value
                lKoz = sheet.cell(row=row_num, column=3).value

                lOMAzon = sheet.cell(row=row_num, column=4).value
                lMegjegyzes = sheet.cell(row=row_num, column=5).value

                if lSorsz is None:
                    break;

                if lSorsz != None or lNev != None or lKoz != None or lOMAzon != None or lMegjegyzes != None:
                    if lSorsz == None: lSorsz = ''

                    if lNev == None: lNev = ''
                    else:            lNev = lNev.replace('\n', ' ').strip()

                    if lKoz == None: lKoz = ''
                    else:            lKoz = lKoz.replace('\n', ' ').strip()

                    if lMegjegyzes == None: lMegjegyzes = ''
                    else:                   lMegjegyzes = lMegjegyzes.replace('\n', ' ').strip()

                    print(lID, lName_Tipus, lName_Ofo, lSorsz, lNev, lKoz, lOMAzon, lMegjegyzes)

                    tanuloChild = root.createElement("tanulo") 
                    tanuloChild.setAttribute('ID', f"{lSorsz}") 
                    tanuloChild.setAttribute('nev', lNev) 
                    tanuloChild.setAttribute('om_azon', f"{lOMAzon}")
                    tanuloChild.setAttribute('megjegyzes', lMegjegyzes)
                    tanuloChild.setAttribute('koz', lKoz)
                    
                    osztalyChild.appendChild(tanuloChild) 
                row_num += 1
    xml_str =  root.toprettyxml(indent="", newl="")
    # xml_str =  root.toprettyxml(indent ="\t")
      

    save_path_file = "iskola.xml"
    # print(xml_str)
    with  open(save_path_file, "w") as f:
        f.write(xml_str)  


    # dom = ET.parse(xml_filename)
    # xslt = ET.parse(xsl_filename)
    # transform = ET.XSLT(xslt)
    # newdom = transform(dom)
    # print(ET.tostring(newdom, pretty_print=True))
    dom = ET.parse("iskola.xml")
    xslt = ET.parse("iskola.xslt")
    transform = ET.XSLT(xslt)
    newdom = transform(dom)
    html = ET.tostring(newdom, pretty_print=True)
    print(html)
    save_path_file = "iskola.html"
        # print(xml_str)
    with  open(save_path_file, "w") as f:
        f.write(str(html))  

        
if __name__ == "__main__":
    main()
