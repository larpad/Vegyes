from openpyxl import load_workbook
import pandas as pd

from iskola import Iskola

osztalyok = Iskola()
path = r"C:\Users\Apu\Desktop\Névsor.xlsx"
path = r"C:\Users\Apu\Desktop\Névsor22.xlsx"

xls = pd.ExcelFile(path)
sheets = xls.sheet_names
#print(sheets)

wb = load_workbook(path, data_only=True)

for sheet in sheets:
    sh = wb[sheet]
    if sheet[0].isdigit():
        szam = int(sheet[0])
        if szam in range(5, 9):
            x = 1
            y = 1
            kilep = False
            osztaly = osztalyok.create_osztaly()
            while not kilep:
                while not kilep:
                    cella = sh.cell(row=x, column=y)
                    if cella.value is not None:
                        osztaly.set(cella.value)
                        kilep = True
                    else:
                        x = x + 1
                if kilep:
                    break
                else:
                    y = y + 1

            x = x + 1
            y = 1
            kilep = False
            while not kilep:
                while not kilep:
                    cella = sh.cell(row=x, column=y)
                    if cella.value is not None:
                        kilep = True
                    else:
                        x = x + 1
                if kilep:
                    break
                else:
                    y = y + 1

            y2 = y + 1
            y3 = y + 3

            x = x + 1
            kilep = False
            while not kilep:
                cella2 = sh.cell(row=x, column=y2)
                cella3 = sh.cell(row=x, column=y3)
                if cella3.value is not None:
                    osztaly.add_tanulo(cella2.value, cella3.value)
                    x = x + 1
                else:
                    break

#print(osztalyok)
#for gyerek in osztalyok.get_list():
#    print(gyerek)

df = pd.DataFrame(osztalyok.get_list())
df.columns[0].set_names("fasd")
print(df)
print(df.info())
